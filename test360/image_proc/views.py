from django.http import HttpResponse
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
import os
from paddleocr import PaddleOCR
import openpyxl
import cv2

# Create your views here.
def process_image(request):
    if request.method == 'POST' and request.FILES['image']:
        # Get the uploaded image
        image = request.FILES['image']
        # Save the image to a temporary location
        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        image_path = fs.path(filename)

        # Initialize the OCR model
        ocr = PaddleOCR(use_angle_cls=True, language='en')

        # Process the image with the OCR model
        image = cv2.imread(image_path)
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
        result = ocr.ocr(image)

        # Create bounding boxes
        result1 = result[0]
        boxes = [line[0] for line in result1]
        txts = [line[1][0] for line in result1]
        scores = [line[1][1] for line in result1]

        for box in boxes:
            p0, p1, p2, p3 = box
            cv2.line(image, (int(p0[0]), int(p0[1])), (int(p1[0]), int(p1[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p1[0]), int(p1[1])), (int(p2[0]), int(p2[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p3[0]), int(p3[1])), (int(p0[0]), int(p0[1])), (255, 0, 0), 2)

        # Save the image with bounding boxes
        processed_image_path = fs.path('processed_image.jpg')
        cv2.imwrite(processed_image_path, image)

        # Create a new Excel workbook and add the results to it
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(result1)):
            sheet.cell(row=i + 1, column=1).value = result1[i][1][0]
            sheet.cell(row=i + 1, column=2).value = result1[i][1][1]

        # Save the Excel file to a temporary location
        excel_filename = os.path.splitext(filename)[0] + '.xlsx'
        excel_path = fs.path(excel_filename)
        workbook.save(excel_path)

        # Get the URL for the processed image and Excel file
        processed_image_url = fs.url('processed_image.jpg')
        excel_url = fs.url(excel_filename)

        # Render the process_image template with the processed image path
        return render(request, 'image_proc/process_image.html', {'processed_image_path': processed_image_path})

    return render(request, 'image_proc/process_image.html')

'''def process_image(request):
    if request.method == 'POST' and request.FILES['image']:
        # Get the uploaded image
        image = request.FILES['image']
        # Save the image to a temporary location
        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        image_path = fs.path(filename)

        # Initialize the OCR model
        ocr = PaddleOCR(use_angle_cls=True, language='en')

        # Process the image with the OCR model
        image = cv2.imread(image_path)
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
        result = ocr.ocr(image)

        # Create bounding boxes
        result1 = result[0]
        boxes = [line[0] for line in result1]
        txts = [line[1][0] for line in result1]
        scores = [line[1][1] for line in result1]

        for box in boxes:
            p0, p1, p2, p3 = box
            cv2.line(image, (int(p0[0]), int(p0[1])), (int(p1[0]), int(p1[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p1[0]), int(p1[1])), (int(p2[0]), int(p2[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p3[0]), int(p3[1])), (int(p0[0]), int(p0[1])), (255, 0, 0), 2)

        # Save the image with bounding boxes
        cv2.imwrite(fs.path('processed_image.jpg'), image)

        # Create a new Excel workbook and add the results to it
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(result1)):
            sheet.cell(row=i + 1, column=1).value = result1[i][1][0]
            sheet.cell(row=i + 1, column=2).value = result1[i][1][1]

        # Save the Excel file to a temporary location
        excel_filename = os.path.splitext(filename)[0] + '.xlsx'
        excel_path = fs.path(excel_filename)
        workbook.save(excel_path)

        # Get the URL for the processed image and Excel file
        processed_image_url = fs.url('processed_image.jpg')
        excel_url = fs.url(excel_filename)

        # Render the results template with the processed image and Excel file URLs
        return render(request, 'image_proc/results.html', {'processed_image_url': processed_image_url, 'excel_url': excel_url})

    return render(request, 'image_proc/process_image.html')'''
