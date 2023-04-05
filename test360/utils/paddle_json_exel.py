#Libraries and paddleOCR
from paddleocr import PaddleOCR, draw_ocr
from matplotlib import pyplot as plt
import cv2
from PIL import Image
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import math
import copy
import json
from shapely.geometry import Polygon


ocr = PaddleOCR(use_angle_cls=True,lang='en')
plt.rcParams['figure.dpi'] = 300

#Load image
name = f"pn3.jpg"
img_path = f"media/{name}"
image = cv2.imread(img_path)
image1 = cv2.imread(img_path)
image = cv2.cvtColor(image,cv2.COLOR_RGB2BGR)
image1 = cv2.cvtColor(image1,cv2.COLOR_RGB2BGR)

#functions
def PadlaOne(image, j=0, i=0):
    result = ocr.ocr(image)
    result1 = result[0]
    boxes = [line[0] for line in result1]
    plt.imsave(f"media/CroppedImages/{j}-{i}.png", image)
    return (result[0])


output_file = {}


def PadlaCrop(image, dx=1000, dy=650, w=1500, h=1000):
    res = []
    lY, lX = image.shape[:2]
    kY = math.ceil(lY / w) + math.ceil((math.ceil(lY / w) * (w - dy) - (math.ceil(lY / w) * w - lY)) / w)
    kX = math.ceil(lX / h) + math.ceil((math.ceil(lX / h) * (h - dx) - (math.ceil(lX / h) * h - lX)) / h)
    x1, y1 = 0, 0
    for j in range(kY):
        x1 = 0
        for i in range(kX):
            crop_img = image[y1:y1 + h, x1:x1 + w]
            r = PadlaOne(crop_img, j, i)
            # print(len(r),r)
            for lis in r:
                li = [x.copy() for x in lis[0]]
                for q in range(4):
                    li[q][0] += i * dx
                    li[q][1] += j * dy
                lis[0] = [lis[0], li]

            res.append([f"{j}-{i}", r])
            x1 = min(x1 + dx, lX)

        y1 = min(y1 + dy, lY)
    print(kY, kX)
    uRes = UniteFrame(res, kY, kX, [w / 2, h / 2])
    DrawBox(image, uRes, output_file)
    return (uRes)
    # return[res]


def calculate_overlap_percentage(A, B):
    p1 = Polygon(A)
    p2 = Polygon(B)
    aP1 = p1.area
    aP2 = p2.area
    p = p1.intersection(p2).area
    return p / aP1 * 100, p / aP2 * 100


def DrawBox(image, res, output_file):
    boxes = []
    for re in res:
        p0, p1, p2, p3 = re[0][1]
        boxes.append([int(p0[0]), int(p0[1]), int(p1[0]), int(p1[1]), int(p2[0]), int(p2[1]), int(p3[0]), int(p3[1])])
        cv2.line(image, (int(p0[0]), int(p0[1])), (int(p1[0]), int(p1[1])), (255, 0, 0), 2)
        cv2.line(image, (int(p1[0]), int(p1[1])), (int(p2[0]), int(p2[1])), (255, 0, 0), 2)
        cv2.line(image, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), (255, 0, 0), 2)
        cv2.line(image, (int(p3[0]), int(p3[1])), (int(p0[0]), int(p0[1])), (255, 0, 0), 2)

        with open('result.json', 'w') as f:
            json.dump(boxes, f)


def RectangleСenter(points):
    xt, yt = 0, 0
    for a in points:
        xt += a[0]
        yt += a[1]
    point = xt / len(points), yt / len(points)
    return (point)


def LengthTo(point1, point2):
    dist = math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)
    return (dist)


def UniteFrame(res, kY, kX, CropImCentr):
    cor = []
    for i in range(kY):
        for j in range(kX):
            k = j + i * kX
            isQ = [0] * len(res[k][1])
            if (j != 0):
                UniteTwoFrame(res[k][1], res[k - 1][1], isQ, CropImCentr)
            if (i != 0):
                kUp = j + (i - 1) * kX
                UniteTwoFrame(res[k][1], res[kUp][1], isQ, CropImCentr)
                if (j != kX - 1):
                    kUpR = j + 1 + (i - 1) * kX
                    UniteTwoFrame(res[k][1], res[kUpR][1], isQ, CropImCentr)
                if (j != 0):
                    kUpL = j - 1 + (i - 1) * kX
                    UniteTwoFrame(res[k][1], res[kUpL][1], isQ, CropImCentr)
            for e in range(len(res[k][1])):
                if (isQ[e] == 0):
                    cor.append(res[k][1][e])
    return (cor)


def UniteTwoFrame(listTh, listUp, listG, CropImCentr):
    d = 30
    for i, a in enumerate(listTh):
        for b in listUp:
            ac = a[0][1]
            bc = b[0][1]
            c, c1 = calculate_overlap_percentage(ac, bc)
            # print(f"SS - {a}\n{b}\n{c}\n\n")
            # print(f"{a[1][0]}-/\-{b[1][0]}-/\-{c}/{c1}\n")
            if (max(c, c1) > 60):
                listG[i] = 1
                if (c < c1):
                    # print(f"{a[0][1]}\n{a[0][0]}\n{a[1]}")
                    for i, v in enumerate(a[0][1]): b[0][1][i] = v
                    for i, v in enumerate(a[0][0]): b[0][0][i] = v
                    b[1] = a[1]
                a = b
                break
    return (listG)


def Angles(res, imageSize, cYaw, cPith):
    li = [{}] * len(res)
    for i, re in enumerate(res):
        a = Angle2(re[0][1], cYaw, cPith, imageSize)
        li[i] = [re[1][0], a]
    return li


def Angle(points, cYaw, cPith, imageSize):
    fx, fy = 2.25, 1.55
    ry, rx = imageSize[0] / 2, imageSize[1] / 2
    centr = RectangleСenter(points)
    dYaw = math.atan(((rx - centr[0]) * math.sin(rx / ry) * math.tan(fx / 2)) / rx)
    dPith = math.atan(((ry - centr[1]) * math.tan(fy / 2)) / ry)
    return (cYaw - dYaw, cPith - dPith)


def Angle2(points, Yaw, Pith, imageSize):
    hfov, vfov = 129, 89
    cy, cx = imageSize[0] / 2, imageSize[1] / 2
    x, y = RectangleСenter(points)
    dYaw = math.atan((x - cx) * math.tan(hfov / 2 * math.pi / 180) / cx)
    dPitch = math.atan((y - cy) * math.tan(vfov / 2 * math.pi / 180) / ((cx / 2) + abs(y - cy)))

    return (Yaw - dYaw, Pith - dPitch)


def Exel(angles):
    file = 'media/exel/Test.xlsx'
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name('List1')

    ps = 58688

    k = 1
    for i, an in enumerate(angles):
        sheet.cell(row=i + 1, column=1).value = an[0]
        sheet.cell(row=i + 1,
                   column=2).value = f"https://preview.r2s.net/app/model/projects/50/visual-model/asset/{ps}?Yaw={an[1][0]}8&Pitch={an[1][1]}8&Roll=0&Zoom=1"
        # sheet.cell(row=k, column=4).value = f"https://preview.r2s.net/app/model/projects/50/visual-model/asset/{ps}?Yaw={1}&Pitch={2}&Roll=0&Zoom=1"

    wb.save("media/exel/Test.xlsx")

#image processing
res = PadlaCrop(image)

image = cv2.cvtColor(image,cv2.COLOR_RGB2BGR)
cv2.imwrite(f"media/proc_{name}", image)