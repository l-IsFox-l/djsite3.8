#model activation
from paddleocr import PaddleOCR, draw_ocr
from matplotlib import pyplot as plt
import cv2
from PIL import Image
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import math

ocr = PaddleOCR(use_angle_cls=True,lang='en')
plt.rcParams['figure.dpi'] = 300

#open image
name = f"test.png"
img_path = f"./Images/{name}"
image = cv2.imread(img_path)
image = cv2.cvtColor(image,cv2.COLOR_RGB2BGR)


# fuctions
def Padleone(image, j, i):
    result = ocr.ocr(image)
    im = image
    result1 = result[0]
    return (result[0])


def PadleCrop(image, dx=650, dy=650, h=1000, w=1500):
    res = []
    lY, lX = image.shape[:2]
    kY = math.ceil(lY / w) + math.ceil((math.ceil(lY / w) * (w - dy) - (math.ceil(lY / w) * w - lY)) / w)
    kX = math.ceil(lX / h) + math.ceil((math.ceil(lX / h) * (h - dx) - (math.ceil(lX / h) * h - lX)) / h)
    x1 = 0
    y1 = 0
    for j in range(kY):
        x1 = 0
        for i in range(kX):
            crop_img = image[y1:y1 + h, x1:x1 + w]
            r = PadlaOne(crop_img, j, i)
            print(len(r), r)
            for li in r:
                for q in range(4):
                    li[0][q][0] = li[0][q][0] + i * dx
                    li[0][q][1] = li[0][q][1] + j * dy

            res.append([f"{j}-{i}", r])
            x1 = min(x1 + dx, lX)

        y1 = min(y1 + dy, lY * 2)
    UniteFrame(res, kY, kX)
    DrawBox(image, res)
    # print(kY,kX)
    return (res)


def UniteFrame(res, kY, kX):
    for i in range(kY):
        for j in range(kX):
            k = j + i * kX
            if (j != 0):
                t = UniteLeftFrame(res[k][1], res[k - 1][1])
                res[k][1] = t[0]
                res[k - 1][1] = t[1]
            if (i != 0):
                kUp = j + (i - 1) * kX
                t1 = UniteUpFrame(res[k][1], res[kUp][1])
                res[k][1] = t1[0]
                res[kUp][1] = t1[1]
            if (i != 0 and j != kX - 1):
                kUpR = j + 1 + (i - 1) * kX
                t1 = UniteUpFrame(res[k][1], res[kUpR][1])
                res[k][1] = t1[0]
                res[kUpR][1] = t1[1]


def UniteLeftFrame(listTh, listPr):
    d = 30
    listC = []
    for a in listTh:
        q = True
        for b in listPr:

            # print(f"\n Left \n{a}\n{b}\n{a[0][0][1]} <> {b[0][0][1]} - {b[0][1][1]} \n {a[0][0][0]} - {b[0][1][0]} \n ")

            if (((a[0][0][1] > b[0][0][1] - d and a[0][0][1] < b[0][1][1] + d) or
                 (a[0][0][1] < b[0][0][1] + d and a[0][0][1] > b[0][1][1] - d)) and
                    a[0][0][0] < b[0][1][0] + d):
                # print("true")
                c = [[min(a[0][0][0], b[0][0][0]), min(a[0][0][1], b[0][0][1])],
                     [max(a[0][1][0], b[0][1][0]), min(a[0][1][1], b[0][1][1])]
                    , [max(a[0][2][0], b[0][2][0]), max(a[0][2][1], b[0][2][1])],
                     [min(a[0][3][0], b[0][3][0]), max(a[0][3][1], b[0][3][1])]]
                ct = [(b[1][0] + " // " + a[1][0]), max(b[1][1], a[1][1])]
                listC.append([c, ct])
                listPr.remove(b)
                q = False
                break
        if (q): listC.append(a)
    return ([listC, listPr])


def UniteUpFrame(listTh, listUp):
    d = 30
    listC = []
    for a in listTh:
        q = True
        for b in listUp:

            # print(f"\n UP \n{a}\n{b}\n{a[0][0][1]} - {b[0][0][1]}\n{a[0][2][1]} - {b[0][2][1]}\n{a[0][0][0]} - {b[0][0][0]}\n{a[0][1][0]} - {b[0][1][0]} \n ")

            if (((a[0][0][1] > b[0][0][1] - d and a[0][0][1] < b[0][0][1] + d) or
                 (a[0][2][1] > b[0][2][1] - d and a[0][2][1] < b[0][2][1] + d)) and
                    ((a[0][0][0] > b[0][0][0] and a[0][0][0] < b[0][1][0] + d) or
                     (a[0][0][0] < b[0][0][0] and a[0][1][0] + d > b[0][1][0]))):
                c = [[min(a[0][0][0], b[0][0][0]), min(a[0][0][1], b[0][0][1])],
                     [max(a[0][1][0], b[0][1][0]), min(a[0][1][1], b[0][1][1])]
                    , [max(a[0][2][0], b[0][2][0]), max(a[0][2][1], b[0][2][1])],
                     [min(a[0][3][0], b[0][3][0]), max(a[0][3][1], b[0][3][1])]]
                # print("true")
                ct = [(b[1][0] + " // " + a[1][0]), max(b[1][1], a[1][1])]
                listC.append([c, ct])
                listUp.remove(b)
                q = False
                break
        if (q): listC.append(a)
    return ([listC, listUp])


def DrawBox(image, res):
    for re in res:
        for r in re[1]:
            p0, p1, p2, p3 = r[0]
            cv2.line(image, (int(p0[0]), int(p0[1])), (int(p1[0]), int(p1[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p1[0]), int(p1[1])), (int(p2[0]), int(p2[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), (255, 0, 0), 2)
            cv2.line(image, (int(p3[0]), int(p3[1])), (int(p0[0]), int(p0[1])), (255, 0, 0), 2)

#functions execution
re = PadleCrop(image)

#ExcelOutPut
file = 'exel/OutPut.xlsx'
wb = load_workbook(file)
sheet = wb.get_sheet_by_name('Sheet1')

ps = 16113

k = 1
for i in range(len(re)):
    for j in range(len(re[i][1])):
        sheet.cell(row=k, column=2).value = re[i][0]
        sheet.cell(row=k, column=3).value = re[i][1][j][1][0]
        #sheet.cell(row=k, column=4).value = f"https://preview.r2s.net/app/model/projects/50/visual-model/asset/{ps}?Yaw={ang[i][j][0]}8&Pitch={ang[i][j][1]}8&Roll=0&Zoom=1"
        sheet.cell(row=k, column=4).value = f"https://preview.r2s.net/app/model/projects/50/visual-model/asset/{ps}?Yaw=448&Pitch=448&Roll=0&Zoom=1"
        k = k + 1

wb.save("exel/OutPut.xlsx")

#end of script (for now)