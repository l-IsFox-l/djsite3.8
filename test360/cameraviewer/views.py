from django.shortcuts import redirect
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import FileSystemStorage

# For PaddleOCR
from paddleocr import PaddleOCR, draw_ocr
from matplotlib import pyplot as plt
import cv2
from PIL import Image
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import math
import copy
from shapely.geometry import Polygon
import py360convert as p36
import json
from PIL import ImageEnhance,ImageFont,ImageDraw

# Create your views here.

def vr_view(request):
        context = {'image_url': '/media/pn_58688.jpg'}
        return render(request, 'cameraviewer/vr_view.html', {'images': context})

def process_panorama(request):
    if request.method == 'POST' and 'process_image' in request.POST:
        # Load model
        ocr = PaddleOCR(use_angle_cls=True, lang='en')
        plt.rcParams['figure.dpi'] = 300

        # Script
        def GetPatch(i=0):
            name = ["media", "Images"]
            patch = [f"{name[0]}", f"{name[0]}/{name[1]}"]
            if (i == -1):
                return patch
            else:
                return (patch[i])

        def CreateFolder():
            patch = GetPatch(-1)
            for pa in patch:
                if (os.path.exists(pa) == False): os.makedirs(pa)

        def PaddleOcr(image):
            result = ocr.ocr(image)
            return (result[0])

        def ImageCrop(image, dx=1000, dy=650, w=1500, h=1000):
            res = []
            lY, lX = image.shape[:2]
            kY = math.ceil(lY / w) + math.ceil((math.ceil(lY / w) * (w - dy) - (math.ceil(lY / w) * w - lY)) / w)
            kX = math.ceil(lX / h) + math.ceil((math.ceil(lX / h) * (h - dx) - (math.ceil(lX / h) * h - lX)) / h)
            x1, y1 = 0, 0
            for j in range(kY):
                x1 = 0
                for i in range(kX):
                    im = image.copy()
                    crop_img = im[y1:y1 + h, x1:x1 + w]
                    r = PaddleOcr(crop_img)
                    for lis in r:
                        li = [x.copy() for x in lis[0]]
                        for q in range(4):
                            li[q][0] += i * dx
                            li[q][1] += j * dy
                        lis[0] = [lis[0], li]

                    res.append([f"{j}-{i}", r])
                    x1 = min(x1 + dx, lX)

                y1 = min(y1 + dy, lY)
            return ([res, [kX, kY], [w, h]])

        def ProcessLargeImage(image):
            res, k, imSize = ImageCrop(image)
            reU = UniteTextImage(res, k, imSize)
            return reU

        def UniteTextImage(res, k, CropImCentr):
            cor = []
            kX, kY = k
            for re in res:
                cor = UniteTwoTextImage(re[1], cor)
            return (cor)

        def UniteTextPano(cordP):
            cor = []
            for cord in cordP:
                cor = UniteTwoTextPano(cord, cor)
            return cor

        def Rectangle(points1, points2):
            p1 = Polygon(points1)
            p2 = Polygon(points2)
            minx, miny, maxx, maxy = p1.bounds
            rect_coords1 = [(minx, miny), (minx, maxy), (maxx, maxy), (maxx, miny)]
            p1R = Polygon(rect_coords1)
            minx, miny, maxx, maxy = p2.bounds
            rect_coords2 = [(minx, miny), (minx, maxy), (maxx, maxy), (maxx, miny)]
            p2R = Polygon(rect_coords2)
            if abs(p1R.area / p1.area) < abs(p2R.area / p2.area):
                return True
            else:
                return False

        def UniteTwoTextPano(listTh, listPr):
            listG = []
            de = []
            for a in listTh:
                q = True
                for b in listPr:
                    c, c1 = CalculateOverlapPercentage(a[1], b[1])
                    if (max(c, c1) > 60):
                        if (len(a[0]) < len(b[0]) or (len(a[0]) == len(b[0]) and Rectangle(a[1], b[1]) == False)):
                            a = copy.deepcopy(b)
                        if b in de:
                            listG.remove(a)
                        else:
                            de.append(b)
                listG.append(a)
            for b in de: listPr.remove(b)
            for b in listPr: listG.append(b)
            return listG

        def UniteTwoTextImage(listTh, listPr):
            listG = []
            de = []
            for a in listTh:
                q = True
                for b in listPr:
                    c, c1 = CalculateOverlapPercentage(a[0][1], b[0][1])
                    if (max(c, c1) > 60):
                        if (len(a[1][0]) < len(b[1][0])):
                            a = copy.deepcopy(b)
                        if b in de:
                            listG.remove(a)
                        else:
                            de.append(b)
                listG.append(a)
            for b in de: listPr.remove(b)
            for b in listPr: listG.append(b)
            return listG

        def CalculateOverlapPercentage(A, B):
            p1 = Polygon(A)
            p2 = Polygon(B)
            aP1 = p1.area
            aP2 = p2.area
            p = p1.intersection(p2).area
            return p / aP1 * 100, p / aP2 * 100

        def RectangleСenter(points):
            xt, yt = 0, 0
            for a in points:
                xt += a[0]
                yt += a[1]
            point = xt / len(points), yt / len(points)
            return (point)

        def LengthTo(point1, point2):
            dist = math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)
            return (dist)

        def DrawBoxImage(image, res):
            for re in res:
                DrawBox(image, re[0][1], (255, 0, 0), 2)

        def DrawBox(image, points, color=(255, 0, 0), lineT=2):
            p0, p1, p2, p3 = points
            cv2.line(image, (int(p0[0]), int(p0[1])), (int(p1[0]), int(p1[1])), color, lineT)
            cv2.line(image, (int(p1[0]), int(p1[1])), (int(p2[0]), int(p2[1])), color, lineT)
            cv2.line(image, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), color, lineT)
            cv2.line(image, (int(p3[0]), int(p3[1])), (int(p0[0]), int(p0[1])), color, lineT)

        def PanoramaCrop(panorama, dYaw=45):
            res = []
            for j in range(1):
                for i in range(math.ceil(360 / dYaw)):
                    yaw = i * dYaw
                    pitch = 0
                    image = p36.e2p(panorama, (145, 145), yaw, pitch, (5000, 5000))
                    res.append([ProcessLargeImage(image), (yaw, pitch)])
            return res

        def CordToAngles4(resP, dYaw, imageSize, fov2):
            angl = []
            for i, re in enumerate(resP):
                an = CordToAngle4(re, imageSize, i * dYaw, 0, fov2)
                angl.append(an)
            return angl

        def CordToAngle4(res, imageSize, cYaw=0, cPitch=0, fov2=(120, 90)):
            li = []
            for re in res[0]:
                a1 = Angle(re[0][1][0], cYaw, cPitch, imageSize, fov2)
                a2 = Angle(re[0][1][1], cYaw, cPitch, imageSize, fov2)
                a3 = Angle(re[0][1][2], cYaw, cPitch, imageSize, fov2)
                a4 = Angle(re[0][1][3], cYaw, cPitch, imageSize, fov2)
                li.append([re[1][0], [a1[:2], a2[:2], a3[:2], a4[:2]]])
            return li

        def Angle(point, yaw, pitch, imageSize, fov2):
            hfov, vfov = fov2
            yaw = math.radians(yaw)
            pitch = math.radians(pitch)
            cy, cx = imageSize[0] / 2, imageSize[1] / 2
            x, y = point
            f1 = cx / math.tan(math.radians(hfov / 2))
            f2 = cy / math.tan(math.radians(vfov / 2))
            dYaw = -1 * math.atan((x - cx) / f1)
            dPitch = math.atan((y - cy) / math.sqrt((x - cx) ** 2 + f2 ** 2))

            yaw -= dYaw
            pitch -= dPitch
            if yaw > 3.14: yaw = (yaw - 6.28)
            if yaw < -3.14: yaw += 6.28

            return (yaw, pitch, 1)

        def Angles4ToCordPano(angles4, panoramSize):
            li = []
            for angle in angles4:
                li.append(Angle4ToCordPano(angle, panoramSize))
            return li

        def Angle4ToCordPano(angle4, panoramSize):
            l = []
            for ang in angle4:
                t = [[]] * 4
                for i in range(4):
                    t[i] = AngleToCordPano(ang[1][i], panoramSize)
                l.append([ang[0], t])
            return l

        def AngleToCordPano(angle, panoramSize):
            vfov = math.radians(180)
            yaw, pitch = angle[0], angle[1] * -1
            x = panoramSize[1] * (yaw + math.pi) / (2 * math.pi)
            y = panoramSize[0] * (pitch + vfov / 2) / vfov
            return ([x, y])

        def CordPanoToAngle4(cordP, panoramSize):
            res = []
            for cord4 in cordP:
                t = []
                for cord in cord4[1]:
                    t.append(CordPanoToAngle(cord, panoramSize))
                res.append([cord4[0], t])
            return res

        def CordPanoToAngles(points, panoramSize):
            res = []
            for point in points:
                p = RectangleСenter(point[1])
                res.append([point[0], CordPanoToAngle(p, panoramSize)])
            return res

        def CordPanoToAngle(point, panoramSize):
            vfov = math.radians(180)
            yaw = 2 * math.pi * point[0] / panoramSize[1] - math.pi
            pitch = (2 * point[1] * vfov - panoramSize[0] * vfov) / (2 * panoramSize[0])
            return (yaw, -1 * pitch)

        def AngleToAngleR2S(angle):
            yaw = (-1 * angle[0]) % 6.283
            pitch = angle[1] + 1.57
            return (yaw, pitch)

        def DrawBoxPano(image, res, s=5):
            for re in res:
                DrawBox(image, re[1], (255, 0, 0), s)

        def ExelR2S(angles, IdCam):
            ps = IdCam
            file = f'media/assets/{ps}/{ps}.xlsx'
            wb = load_workbook(file)
            sheet = wb.active
            for i, an in enumerate(angles):
                angl = AngleToAngleR2S(an[1])
                sheet.cell(row=i + 1, column=1).value = an[0]
                sheet.cell(row=i + 1,
                           column=2).value = f"https://preview.r2s.net/app/model/projects/50/visual-model/asset/{ps}?Yaw={angl[0]}8&Pitch={angl[1]}8&Roll=0&Zoom={1}"
            wb.save(f"media/assets/{ps}/{ps}.xlsx")

        def WriteAllTextPano(cords, panorama):
            image = Image.fromarray(panorama.astype('uint8'), 'RGB')
            boxs = []
            for cord in cords:
                WriteTextImage(cord, image, boxs)
            return image

        def WriteTextImage(cord, image, boxs):
            text, points = cord

            dx = abs(points[0][0] - points[1][0])
            fz = 35
            textSize = 0
            while True:
                textSize = GetPilTextSize(text, fz, "ArialRegular.ttf")
                if (textSize[0] <= dx): break
                fz -= 1
                if (fz <= 20):
                    fz = 20
                    break

            font = ImageFont.truetype("ArialRegular.ttf", fz)
            imText = Image.new('RGB', textSize, (255, 255, 255))
            drawer = ImageDraw.Draw(imText)
            drawer.text((0, 0), text, font=font, fill='Black')
            dy, dx = 0, 0

            xmin, ymin = int(points[2][0]) + 5, int(points[2][1] - textSize[1])
            xmax, ymax = int(xmin + textSize[0]), int(ymin + textSize[1])
            p = Polygon([[xmin, ymin], [xmin, ymax], [xmax, ymax], [xmax, ymin]])

            q = False
            for b in boxs:
                if max(CalculateOverlapPercentage(p, b)) > 5:
                    q = True
                    if (p.bounds[3] > b.bounds[3]):
                        dy = -(p.bounds[1] - b.bounds[3])
                    else:
                        dy = b.bounds[1] - p.bounds[3]
                    break

            p = Polygon([[xmin, ymin + dy], [xmin, ymax + dy], [xmax, ymax + dy], [xmax, ymin + dy]])
            if q:
                for b in boxs:
                    if (max(CalculateOverlapPercentage(p, b)) > 5):
                        dx = abs(b.bounds[0] - p.bounds[2])
                        break
                p = Polygon(
                    [[xmin + dx, ymin + dy], [xmin + dx, ymax + dy], [xmax + dx, ymax + dy], [xmax + dx, ymin + dy]])

            boxs.append(p)
            image.paste(imText, (int(points[2][0]) + 5, int(points[2][1] - textSize[1] + dy)))

        def GetPilTextSize(text, font_size, font_name):
            font = ImageFont.truetype(font_name, font_size)
            size = font.getsize(text)
            return size

        # Load panorama
        IdPan = 58688
        name = f"pn_58688.jpg"
        img_path = f"media/assets/{IdPan}/{name}"
        panorama = cv2.imread(img_path)
        panorama = cv2.cvtColor(panorama, cv2.COLOR_RGB2BGR)
        panorama = np.array(panorama)

        # Functions execution
        re = PanoramaCrop(panorama)

        # Post pr
        angle = CordToAngles4(re, 45, (5000, 5000), (145, 145))  # преоброзовать все координаты в углы 4x
        cordP = Angles4ToCordPano(angle, panorama.shape[:2])  # углы в координаты на панораме 4x
        cord = copy.deepcopy(cordP)
        cordPU = UniteTextPano(cord)  # удаление дубликатов

        DrawBoxPano(panorama, cordPU, 4)
        pano = cv2.cvtColor(panorama, cv2.COLOR_RGB2BGR)
        cv2.imwrite(f"media/assets/{IdPan}/boxes_only_{name}", pano) #just boxes

        anglP = CordPanoToAngles(cordPU, panorama.shape[:2])  # координаты панорамы в углы
        ExelR2S(anglP, IdPan)  # запись в эксэль

        angle4P = CordPanoToAngle4(cordPU, panorama.shape[:2])  # преоброзование координат панорамы в углы 4x
        cordS = Angle4ToCordPano(angle4P, panorama.shape[:2])  # углы в координаты меньшей панорамы 4x
        DrawBoxPano(panorama, cordPU, 2)  # отрисовка рамок

        SmalPanoIm = WriteAllTextPano(cordPU, panorama)  # отрисовка текста
        SmalPanoIm.save(f"media/assets/{IdPan}/proc_{name}")

        # Return response to user
        return redirect('vr_view')